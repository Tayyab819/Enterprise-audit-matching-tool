"""
utils.py — Shared utilities for Internal Audit Makes Easy.

Covers:
  - Excel workbook styling helpers
  - Amount parsing (handles commas, currency symbols, brackets)
  - Text normalisation for fuzzy matching
  - File / DataFrame validation
  - Temporary file management
"""

from __future__ import annotations

import logging
import re
import tempfile
import traceback
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from constants import (
    COL_MAX_WIDTH,
    COL_MIN_WIDTH,
    FILL_WHITE,
    FILL_GREY_EVEN,
    FILL_YELLOW_HI,
    NOISE_WORDS,
)

log = logging.getLogger(__name__)

# ──────────────────────────────────────────────────────────────────────────────
# Logging
# ──────────────────────────────────────────────────────────────────────────────

def setup_logging() -> None:
    """Configure root logger with timestamp, level, and full traceback support."""
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(name)s — %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )


def log_exception(msg: str, exc: Exception) -> None:
    """Log an exception with full stack trace for easier debugging."""
    log.error("%s\n%s", msg, traceback.format_exc())


# ──────────────────────────────────────────────────────────────────────────────
# Amount parsing
# ──────────────────────────────────────────────────────────────────────────────

_CURRENCY_RE   = re.compile(r"[$£€₹¥,\s]")
_BRACKETS_RE   = re.compile(r"\((.+?)\)")


def parse_amount(series: pd.Series) -> pd.Series:
    """
    Robustly convert a mixed-type series to float amounts.

    Handles:
      - Currency symbols  ($, £, €, ₹, ¥)
      - Thousands commas
      - Bracketed negatives  (1,234.56) → -1234.56
      - NaN / blanks → 0.0
    """
    cleaned = (
        series.fillna(0)
              .astype(str)
              .str.strip()
              .str.replace(_CURRENCY_RE, "", regex=True)
              .str.replace(_BRACKETS_RE, r"-\1", regex=True)
    )
    return pd.to_numeric(cleaned, errors="coerce").fillna(0.0)


# ──────────────────────────────────────────────────────────────────────────────
# Text normalisation
# ──────────────────────────────────────────────────────────────────────────────

_MULTI_SPACE_RE = re.compile(r"\s+")
_NON_ALPHA_RE   = re.compile(r"[^a-z0-9\s]")
_NOISE_PATTERN  = re.compile(
    r"\b(?:" + "|".join(re.escape(w) for w in NOISE_WORDS) + r")\b",
    re.IGNORECASE,
)


def normalise_text(value: str) -> str:
    """
    Normalise a string for fuzzy comparison:
      1. Lowercase
      2. Remove punctuation / special chars
      3. Strip common noise words (Ltd, Inc, Co …)
      4. Collapse whitespace
    """
    v = value.strip().lower()
    v = _NON_ALPHA_RE.sub(" ", v)
    v = _NOISE_PATTERN.sub(" ", v)
    v = _MULTI_SPACE_RE.sub(" ", v).strip()
    return v


def normalise_series(series: pd.Series) -> pd.Series:
    """Apply normalise_text to an entire Series efficiently."""
    return series.fillna("").astype(str).apply(normalise_text)


# ──────────────────────────────────────────────────────────────────────────────
# File / DataFrame validation
# ──────────────────────────────────────────────────────────────────────────────

class ValidationError(ValueError):
    """Raised when uploaded data fails validation."""


def validate_dataframe(
    df: pd.DataFrame,
    required_cols: Optional[list[str]] = None,
    label: str = "file",
) -> None:
    """
    Validate a DataFrame before processing.

    Args:
        df: DataFrame to validate.
        required_cols: Column names that must be present.
        label: Human-readable label used in error messages.

    Raises:
        ValidationError: With a descriptive message on failure.
    """
    if df is None:
        raise ValidationError(f"{label}: no data loaded.")
    if df.empty:
        raise ValidationError(f"{label}: file is empty.")
    if required_cols:
        missing = [c for c in required_cols if c not in df.columns]
        if missing:
            raise ValidationError(
                f"{label}: missing required columns: {', '.join(missing)}. "
                f"Available columns: {', '.join(df.columns.tolist())}"
            )


def load_excel(path: str, label: str = "file") -> tuple[pd.DataFrame, str]:
    """
    Load an Excel file into a DataFrame with sanitised column names.

    Returns:
        (DataFrame, warning_message)  — warning is empty string if no issues.

    Raises:
        ValidationError: on read failure or empty file.
    """
    try:
        df = pd.read_excel(path)
    except Exception as exc:
        log_exception(f"Failed to read {label}", exc)
        raise ValidationError(f"Cannot read {label}: {exc}") from exc

    df.columns = df.columns.astype(str).str.strip()

    # Drop fully-empty columns/rows
    df = df.dropna(axis=1, how="all").dropna(axis=0, how="all").reset_index(drop=True)

    if df.empty:
        raise ValidationError(f"{label}: file contains no usable data after cleaning.")

    warning = ""
    from constants import MAX_FILE_ROWS, LARGE_FILE_ROWS
    if len(df) > MAX_FILE_ROWS:
        warning = (
            f"⚠️ {label} has {len(df):,} rows — exceeds the {MAX_FILE_ROWS:,}-row limit. "
            "Please split your file into smaller batches."
        )
        raise ValidationError(warning)
    if len(df) > LARGE_FILE_ROWS:
        warning = (
            f"ℹ️ {label} has {len(df):,} rows — processing may take a few seconds."
        )

    return df, warning


def sanitise_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """
    Standardise string columns: strip whitespace, unify line endings.
    Numeric columns are left untouched.
    """
    for col in df.select_dtypes(include="object").columns:
        df[col] = df[col].astype(str).str.replace(r"[\r\n\t]", " ", regex=True).str.strip()
    return df


# ──────────────────────────────────────────────────────────────────────────────
# Temporary file management
# ──────────────────────────────────────────────────────────────────────────────

_TEMP_FILES: list[Path] = []


def make_temp_xlsx(prefix: str = "audit_") -> str:
    """
    Create a named temporary .xlsx file and register it for cleanup.

    Returns:
        Absolute path string.
    """
    tmp = tempfile.NamedTemporaryFile(
        suffix=".xlsx", prefix=prefix, delete=False
    )
    tmp.close()
    path = Path(tmp.name)
    _TEMP_FILES.append(path)
    return str(path)


def cleanup_temp_files() -> None:
    """Delete all temporary files created during this session."""
    for p in _TEMP_FILES:
        try:
            if p.exists():
                p.unlink()
        except OSError:
            pass
    _TEMP_FILES.clear()


# ──────────────────────────────────────────────────────────────────────────────
# Excel styling — reusable helpers
# ──────────────────────────────────────────────────────────────────────────────

def _thin_border(color: str = "000000") -> Border:
    side = Side(border_style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def style_header(ws: Worksheet, row: int, bg: str = "2F5496") -> None:
    """Apply bold white header styling to a worksheet row."""
    border = _thin_border()
    for cell in ws[row]:
        cell.font      = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
        cell.fill      = PatternFill("solid", start_color=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border


def style_zebra(
    ws: Worksheet,
    start: int,
    end: int,
    even_fill: str = FILL_WHITE,
    odd_fill: str = FILL_GREY_EVEN,
) -> None:
    """Apply alternating row fills and light borders from *start* to *end*."""
    border = _thin_border("D0D0D0")
    for row in ws.iter_rows(min_row=start, max_row=end):
        fill = even_fill if row[0].row % 2 == 0 else odd_fill
        for cell in row:
            cell.fill      = PatternFill("solid", start_color=fill)
            cell.border    = border
            cell.font      = Font(name="Calibri", size=10)
            cell.alignment = Alignment(vertical="center")


def auto_col_width(ws: Worksheet) -> None:
    """Set each column's width based on its longest value."""
    for col_cells in ws.columns:
        width = max(
            (len(str(c.value)) if c.value is not None else 0) for c in col_cells
        )
        letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[letter].width = min(
            max(width + 3, COL_MIN_WIDTH), COL_MAX_WIDTH
        )


def style_sheet(
    ws: Worksheet,
    header_bg: str = "2F5496",
    even_fill: str = FILL_WHITE,
) -> None:
    """One-call helper: header row + zebra rows + column widths + freeze pane."""
    style_header(ws, 1, header_bg)
    if ws.max_row > 1:
        style_zebra(ws, 2, ws.max_row, even_fill)
    auto_col_width(ws)
    ws.freeze_panes = "A2"


def write_summary_section(
    ws: Worksheet,
    row: int,
    title: str,
    color: str,
    kv_pairs: list[tuple[str, object]],
    num_cols: int = 2,
) -> int:
    """
    Write a titled section with key-value rows onto a summary worksheet.

    Args:
        ws: Target worksheet.
        row: Starting row number.
        title: Section header text.
        color: Hex background colour for the section header.
        kv_pairs: List of (label, value) tuples.
        num_cols: Number of columns to merge for the section header.

    Returns:
        The next available row number after this section.
    """
    thin = Side(border_style="thin", color="BDC3C7")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Section title row
    cell = ws.cell(row=row, column=1, value=title)
    cell.font      = Font(bold=True, color="FFFFFF", name="Calibri", size=12)
    cell.fill      = PatternFill("solid", start_color=color)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    ws.row_dimensions[row].height = 26
    row += 1

    # Key-value rows
    for label, value in kv_pairs:
        lc = ws.cell(row=row, column=1, value=label)
        vc = ws.cell(row=row, column=2, value=value)
        bg = FILL_YELLOW_HI if isinstance(value, float) else (
            FILL_GREY_EVEN if row % 2 == 0 else FILL_WHITE
        )
        for c in (lc, vc):
            c.fill      = PatternFill("solid", start_color=bg)
            c.border    = brd
            c.alignment = Alignment(vertical="center", indent=1)
            c.font      = Font(name="Calibri", size=11)
        lc.font = Font(bold=True, name="Calibri", size=11)
        ws.row_dimensions[row].height = 20
        row += 1

    return row
